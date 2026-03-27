from openpyxl import load_workbook
import pandas as pd
import json

def print_excel_preview(input_path):
    wb = load_workbook(input_path, read_only=True)
    ws = wb.active
    print("[PREVIEW] First 10 rows and columns:")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(row)
        if i >= 9:
            break

def find_header_row(input_path):
    wb = load_workbook(input_path, read_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row is None:
            continue
        row_str = [str(cell).lower() if cell else "" for cell in row]
        if any("json" in cell or "api" in cell or "facade" in cell for cell in row_str):
            print(f"[PREVIEW] Header row found at Excel row {i+1}: {row}")
            return i, row
    print("[PREVIEW] No header row with 'json', 'api', or 'facade' found in first 20 rows.")
    return None, None
"""
Export enriched Excel files with normalized descriptions for fields, arrays, and objects.
- Loads an Excel file
- Detects fields, arrays, and objects
- Generates normalized descriptions
- Writes a new Excel file with added description columns and a summary sheet
"""
import os
import sys
from typing import List
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
# Ensure backend is in sys.path for absolute import
import pathlib
BACKEND_DIR = pathlib.Path(__file__).resolve().parents[2]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))
from components.mapping_agents.parsers import parse_excel, ExcelObjectDescriptor



from components.mapping_agents.parsers import _descriptor_is_array

def get_parent_array(path: str) -> str:
    # Returns the parent array path if present, else ""
    # E.g. "root.items[].field" -> "root.items[]"
    if not path:
        return ""
    # Find last array marker
    arr_idx = path.rfind("[")
    dotarr_idx = path.rfind(".[]")
    cut_idx = max(arr_idx, dotarr_idx)
    if cut_idx == -1:
        return ""
    # Find end of array marker
    if path[cut_idx:cut_idx+3] == ".[]":
        return path[:cut_idx+3]
    elif path[cut_idx] == "[":
        # Find matching ]
        close_idx = path.find("]", cut_idx)
        if close_idx != -1:
            return path[:close_idx+1]
        else:
            return path[:cut_idx+1]
    return ""

def summarize_excel_for_llm(input_path: str, output_path: str):
    print_excel_preview(input_path)
    header_info = find_header_row(input_path)
    header_idx = header_info[0] if header_info and isinstance(header_info, tuple) else None
    print(f"[DEBUG] Reading input file: {input_path}")
    print(f"[DEBUG] Input directory: {os.path.dirname(input_path)}")

    # Read with detected header row if available, else try header=0 then header=1
    if header_idx is not None:
        # Read raw without header, then set header from detected row to handle merged/blank cells
        raw = pd.read_excel(input_path, header=None)
        print("[DEBUG] Raw shape:", raw.shape)
        print("[DEBUG] Raw head (first 5 rows):\n", raw.head().to_string())
        # Ensure header_idx is within bounds
        if header_idx < 0 or header_idx >= len(raw):
            raw = pd.read_excel(input_path, header=0)
            df = raw.loc[:, ~raw.columns.str.contains('^Unnamed')]
            print(f"[DEBUG] Detected header row out of bounds; fell back to header=0. Columns: {list(df.columns)}")
        else:
            header_row = raw.iloc[header_idx].fillna("").astype(str).tolist()
            print(f"[DEBUG] Header row values from pandas raw: {header_row}")
            raw.columns = header_row
            df = raw.iloc[header_idx + 1 :].reset_index(drop=True)
            # Drop columns where the header is empty string
            df = df.loc[:, [c for c in df.columns if c and str(c).strip() != ""]]
            print(f"[DEBUG] Columns found using detected header (header={header_idx}):", list(df.columns))
    else:
        df = pd.read_excel(input_path, header=0)
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        print("[DEBUG] Columns found in input Excel (header=0, cleaned):", list(df.columns))
        if not any('json' in c.lower() for c in df.columns):
            df2 = pd.read_excel(input_path, header=1)
            df2 = df2.loc[:, ~df2.columns.str.contains('^Unnamed')]
            print("[DEBUG] Columns found in input Excel (header=1, cleaned):", list(df2.columns))
            if any('json' in c.lower() for c in df2.columns):
                df = df2
    print("[DEBUG] Final columns used:", list(df.columns))
    def test_agentic_excel_summary():
        # This test will run the summarizer on a known test file and print the output file path
        import tempfile
        test_input = "C:/dev/snowchat/backend/test_excels/problem_file.xlsx"
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            test_output = tmp.name
        try:
            summarize_excel_for_llm(test_input, test_output)
            print(f"[AGENTIC TEST] Output written to: {test_output}")
        except Exception as e:
            print(f"[AGENTIC TEST] Error: {e}")
    # Use the columns that contain the JSON path, API, and Façade

    col_map = {}
    for col in df.columns:
        col_lower = col.lower().replace(" ", "")
        if 'json' in col_lower:
            if 'json' not in col_map:
                col_map['json'] = col
        if 'api' in col_lower:
            if 'api' not in col_map:
                col_map['api'] = col
        if 'façade' in col_lower or 'facade' in col_lower:
            if 'facade' not in col_map:
                col_map['facade'] = col

    if 'json' not in col_map:
        print(f"[DEBUG] pandas did not match json column. Columns: {list(df.columns)}")
        # Fallback: use openpyxl to build rows from detected header row
        wb = load_workbook(input_path, read_only=True)
        ws = wb.active
        if header_idx is None:
            raise ValueError(f"No column containing 'json' found in input Excel. Columns found: {list(df.columns)} | Matched: {col_map}")
        # openpyxl rows are 1-based
        header_cells = list(ws[header_idx + 1])
        headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]
        print(f"[DEBUG] openpyxl detected headers: {headers}")
        # build list of dicts from rows below header
        data_rows = []
        for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
            row_dict = {}
            for idx_col, header in enumerate(headers):
                if not header or header.strip() == "":
                    continue
                value = row[idx_col] if idx_col < len(row) else None
                row_dict[header] = value
            data_rows.append(row_dict)
        print(f"[DEBUG] Built {len(data_rows)} data rows from openpyxl fallback")
        # convert to DataFrame
        df = pd.DataFrame(data_rows)
        print(f"[DEBUG] Columns after openpyxl fallback: {list(df.columns)}")
        # rebuild col_map
        col_map = {}
        for col in df.columns:
            col_lower = col.lower().replace(" ", "")
            if 'json' in col_lower and 'json' not in col_map:
                col_map['json'] = col
            if 'api' in col_lower and 'api' not in col_map:
                col_map['api'] = col
            if ('façade' in col.lower() or 'facade' in col.lower()) and 'facade' not in col_map:
                col_map['facade'] = col
        if 'json' not in col_map:
            raise ValueError(f"No column containing 'json' found after openpyxl fallback. Columns found: {list(df.columns)} | Matched: {col_map}")
        # Also create a normalized temp Excel so pandas can read headers properly
        try:
            import tempfile
            from openpyxl import Workbook as XLWorkbook
            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            tmp_path = tmp.name
            tmp.close()
            wb2 = XLWorkbook()
            ws2 = wb2.active
            # write header
            for c_idx, h in enumerate(headers, start=1):
                if h and str(h).strip():
                    ws2.cell(row=1, column=c_idx, value=h)
            # write data rows
            for r_idx, row in enumerate(data_rows, start=2):
                for c_idx, h in enumerate(headers, start=1):
                    if not h or str(h).strip() == "":
                        continue
                    ws2.cell(row=r_idx, column=c_idx, value=row.get(h))
            wb2.save(tmp_path)
            print(f"[DEBUG] Wrote normalized temp Excel for pandas: {tmp_path}")
            # read with pandas from normalized file
            df = pd.read_excel(tmp_path)
            # cleanup temp file
            try:
                os.remove(tmp_path)
            except Exception:
                pass
            print(f"[DEBUG] pandas columns after reading normalized file: {list(df.columns)}")
        except Exception as exc:
            print(f"[DEBUG] Failed to create normalized excel for pandas: {exc}")

    def get_parent(path):
        if not isinstance(path, str) or "." not in path:
            return ""
        parent = path.rsplit(".", 1)[0]
        if parent.endswith("[]"):
            parent = parent[:-2]
        return parent

    array_fields = df[df[col_map['json']].astype(str).str.contains(r"\[\]")]
    array_roots = set(array_fields[col_map['json']].apply(lambda p: p.split("[")[0] if "[" in p else p))

    rows = []
    for idx, row in df.iterrows():
        path = row.get(col_map['json'], "")
        if not isinstance(path, str) or not path.strip():
            continue
        parent_array = get_parent(path)
        field_name = path.split(".")[-1].replace("[]", "")
        # Classify type
        if any(path.startswith(ar) for ar in array_roots):
            if path in array_roots:
                type_ = "Array"
            else:
                type_ = "Array Field"
        elif parent_array:
            type_ = "Dynamic Field"
        else:
            type_ = "Static Field"
        out_row = {
            "Type": type_,
            "JSON Path": path,
            "Field Name": field_name,
            "Parent Array": parent_array,
            "LLM Description": row.get("Description", ""),
        }
        # Add API and Façade columns if present
        if 'api' in col_map:
            out_row['API'] = row.get(col_map['api'], "")
        if 'facade' in col_map:
            out_row['Façade'] = row.get(col_map['facade'], "")
        rows.append(out_row)

    # Deduplicate rows by Field Name + Parent Array + API + Façade
    import re

    def normalize_path(p: str) -> str:
        if not isinstance(p, str):
            return p
        # replace numeric indices [0], [1] with [] for normalization
        return re.sub(r"\[\d+\]", "[]", p)

    # Build set of array roots from all paths (normalized)
    normalized_paths = [normalize_path(r.get('JSON Path', '')) for r in rows]
    array_roots = set()
    for np in normalized_paths:
        if not np:
            continue
        parts = np.split('.')
        acc = []
        for part in parts:
            acc.append(part)
            if '[]' in part:
                array_roots.add('.'.join(acc))

    # Deduplicate
    dedupe = {}
    for r in rows:
        api = r.get('API', '')
        facade = r.get('Façade', r.get('Facade', ''))
        key = (r['Field Name'], r['Parent Array'], api, facade)
        if key in dedupe:
            # merge LLM Description if missing
            existing = dedupe[key]
            if not existing.get('LLM Description') and r.get('LLM Description'):
                existing['LLM Description'] = r.get('LLM Description')
        else:
            dedupe[key] = r.copy()
    deduped_rows = list(dedupe.values())

    # Build hierarchical array summaries
    # Find nested arrays: any array_root that is a prefix of another
    def is_prefix(a, b):
        return b.startswith(a + '.') or b == a

    array_roots_sorted = sorted(array_roots, key=lambda x: x.count('.'))
    array_summaries = []
    for root in sorted(array_roots, key=lambda x: x):
        fields = sorted({r['Field Name'] for r in deduped_rows if r['Parent Array'] == root})
        subarrays = sorted([a for a in array_roots if a != root and is_prefix(root, a)])
        # find API/Facade for this array if any row has it
        api = ''
        facade = ''
        for r in deduped_rows:
            if r['Parent Array'] == root or r['JSON Path'] == root:
                api = api or r.get('API', '')
                facade = facade or r.get('Façade', r.get('Facade', ''))
        array_summaries.append({
            'Array Root': root,
            'Parent Array': get_parent(root),
            'Fields': ', '.join(fields),
            'Subarrays': ', '.join(subarrays),
            'API': api,
            'Façade': facade,
        })

    # Final deduped rows (make sure dedupe is applied before any final writes)
    out_df = pd.DataFrame(deduped_rows)
    # ensure column order
    cols = [c for c in ['Type', 'JSON Path', 'Field Name', 'Parent Array', 'LLM Description', 'API', 'Façade'] if c in out_df.columns]
    out_df = out_df[cols]
    summary_df = pd.DataFrame(array_summaries)
    # Write Excel with two sheets from deduped rows
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        out_df.to_excel(writer, index=False, sheet_name='LLM_Input')
        summary_df.to_excel(writer, index=False, sheet_name='Array_Summaries')
    print(f"Exported deduplicated LLM-ready summary + arrays to {output_path}")

    # Also produce a JSON file using JSON Path as the key for each deduped row
    try:
        import json
        json_out = {}
        for r in deduped_rows:
            jp = r.get('JSON Path') or r.get('JSON_Path') or r.get('json_path')
            if not jp:
                # skip rows without JSON path
                continue
            # Normalize path indices to [] form
            jp_norm = normalize_path(jp)
            # Represent the row as a dict without duplicating large objects
            json_out[jp_norm] = {
                'Type': r.get('Type', ''),
                'Field Name': r.get('Field Name', ''),
                'Parent Array': r.get('Parent Array', ''),
                'LLM Description': r.get('LLM Description', ''),
                'API': r.get('API', ''),
                'Façade': r.get('Façade', r.get('Facade', '')),
            }
        json_path = os.path.splitext(output_path)[0] + '.json'
        with open(json_path, 'w', encoding='utf-8') as jf:
            json.dump(json_out, jf, ensure_ascii=False, indent=2)
        print(f"Exported deduplicated JSON representation to {json_path}")
    except Exception as e:
        print(f"[WARN] Failed to write JSON output: {e}")

    # Build final single-sheet rows from deduped rows and add array summary rows
    final_rows = [r.copy() for r in deduped_rows]
    for array_root in sorted(array_roots):
        field_names = sorted(set(r['Field Name'] for r in deduped_rows if r['Parent Array'] == array_root and r['Type'] == 'Array Field'))
        if field_names:
            out_row = {
                'Type': 'Array Summary',
                'JSON Path': array_root,
                'Field Name': ', '.join(field_names),
                'Parent Array': get_parent(array_root),
                'LLM Description': '',
            }
            if 'api' in col_map:
                out_row['API'] = ''
            if 'facade' in col_map:
                out_row['Façade'] = ''
            final_rows.append(out_row)

    # Write to single-sheet Excel with explicit column order (deduped)
    base_cols = ['Type', 'JSON Path', 'Field Name', 'Parent Array', 'LLM Description']
    if 'api' in col_map:
        base_cols.append('API')
    if 'facade' in col_map:
        base_cols.append('Façade')
    type_order = {'Static Field': 0, 'Dynamic Field': 1, 'Array Field': 2, 'Array': 3, 'Array Summary': 4}
    out_df = pd.DataFrame(final_rows, columns=base_cols)
    out_df['_type_order'] = out_df['Type'].map(type_order)
    out_df = out_df.sort_values(['_type_order', 'JSON Path']).reset_index(drop=True)
    out_df = out_df.drop(columns=['_type_order'])
    # Overwrite the Excel single-sheet (keeps Array_Summaries sheet separate if needed)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        out_df.to_excel(writer, index=False, sheet_name='LLM_Input')
        # preserve array summaries sheet
        summary_df.to_excel(writer, index=False, sheet_name='Array_Summaries')


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python export_excel_with_descriptions.py <input.xlsx> <output.xlsx>")
        sys.exit(1)
    summarize_excel_for_llm(sys.argv[1], sys.argv[2])
    print(f"Exported LLM-ready summary to {sys.argv[2]}")

def regenerate_and_inspect(input_path: str, output_path: str):
    """Run summarizer and return some stats as a dict (for programmatic inspection)."""
    summarize_excel_for_llm(input_path, output_path)
    # read back the generated Excel and JSON
    stats = {}
    try:
        df = pd.read_excel(output_path, sheet_name='LLM_Input')
        stats['llm_rows'] = len(df)
    except Exception as e:
        stats['llm_error'] = str(e)
    json_p = os.path.splitext(output_path)[0] + '.json'
    if os.path.exists(json_p):
        try:
            with open(json_p, 'r', encoding='utf-8') as jf:
                j = json.load(jf)
            stats['json_entries'] = len(j)
        except Exception as e:
            stats['json_error'] = str(e)
    else:
        stats['json_error'] = 'notfound'
    return stats

